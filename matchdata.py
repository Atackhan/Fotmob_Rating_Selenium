import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook

#  WebDriver Ayarları
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)

#  Excel dosyasını oluştur veya yükle
excel_file = "fotmob_data1.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Maç Tarihi", "Ev Sahibi", "Deplasman", "Oyuncu Adı", "Oyuncu Rating"])

try:
    #  Sayfa sayısı kadar döngü
    for page_num in range(7):  # 0'dan 6'ya kadar (7 sayfa)
        url = f"https://www.fotmob.com/tr/leagues/47/matches/premier-league?season=2023-2024&group=by-date&page={page_num}"
        driver.get(url)
        time.sleep(2)

        #  Sayfa biraz kaydırılarak içeriklerin yüklenmesi sağlanıyor
        driver.execute_script("window.scrollBy(0, 200);")
        time.sleep(0.5)

        print(f" {page_num}. sayfa işleniyor...")

        x = 1
        while True:
            y = 1
            while True:
                try:
                    match_xpath = f"/html/body/div[1]/main/main/section/div/div[2]/section/div[2]/div[2]/section[{x}]/a[{y}]"
                    match_element = driver.find_element(By.XPATH, match_xpath)

                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", match_element)
                    time.sleep(1)
                    match_element.click()
                    time.sleep(4)

                    #  Yeni XPath ile istatistik sekmesine tıklama
                    try:
                        stats_button = driver.find_element(By.XPATH, "/html/body/div[1]/main/main/div[2]/div/div[1]/div[1]/div/nav/button[5]")
                        stats_button.click()
                        time.sleep(3)
                    except:
                        print(" Hata: İstatistik butonu tıklanamadı, atlanıyor.")
                        driver.back()
                        time.sleep(4)
                        y += 1
                        continue

                    #  Maç tarihi
                    try:
                        match_date = driver.find_element(By.XPATH, "/html/body/div[1]/main/main/div[2]/div/div[1]/div[1]/div/section/div/div[3]/section/ul/li[1]/div/time/span[1]").text
                    except:
                        match_date = "Bilinmiyor"

                    #  Takım isimleri
                    try:
                        home_team = driver.find_element(By.XPATH, "/html/body/div[1]/main/main/div[2]/div/div[1]/div[1]/div/section/div/section/header/div[1]/a/div/span/span[2]").text
                    except:
                        home_team = "Bilinmiyor"
                    try:
                        away_team = driver.find_element(By.XPATH, "/html/body/div[1]/main/main/div[2]/div/div[1]/div[1]/div/section/div/section/header/div[3]/a/div/span").text
                    except:
                        away_team = "Bilinmiyor"

                    #  Oyuncu verileri
                    z = 1
                    while True:
                        try:
                            player_name_xpath = f"/html/body/div[1]/main/main/div[2]/div/div[1]/div[3]/div/div[2]/table/tbody/tr[{z}]/td[1]/div/a/div/span"
                            player_rating_xpath = f"/html/body/div[1]/main/main/div[2]/div/div[1]/div[3]/div/div[2]/table/tbody/tr[{z}]/td[2]/div/div/span"

                            player_name = driver.find_element(By.XPATH, player_name_xpath).text
                            player_rating = driver.find_element(By.XPATH, player_rating_xpath).text

                            ws.append([match_date, home_team, away_team, player_name, player_rating])
                            z += 1
                        except:
                            break  # Oyuncular bitti

                    wb.save(excel_file)
                    print(f"✔ Maç Kaydedildi: {match_date} | {home_team} vs {away_team}")

                    driver.back()
                    time.sleep(1.5)
                    y += 1

                except:
                    break  # section[x] altındaki maçlar bitti

            try:
                driver.find_element(By.XPATH, f"/html/body/div[1]/main/main/section/div/div[2]/section/div[2]/div[2]/section[{x+1}]")
                x += 1
            except:
                break  # Tüm sectionlar bitti

        print(f" {page_num}. sayfa tamamlandı. Bir sonraki sayfaya geçiliyor...")

except KeyboardInterrupt:
    print("Program kullanıcı tarafından durduruldu. Şimdiye kadar toplanan veriler kaydediliyor...")

except Exception as e:
    print(f"Beklenmeyen bir hata oluştu: {e}")

finally:
    #  Excel dosyasını güvenli şekilde kaydet ve kapat
    wb.save(excel_file)
    wb.close()
    driver.quit()
    print(" Veriler başarıyla kaydedildi. Tarayıcı kapatıldı.")
